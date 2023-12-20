import calendar
import datetime
import os

import openpyxl
import telebot
from django import setup
from django.db.models import Q
from django.utils import timezone
from telebot import types
from datetime import datetime, timedelta

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "TaskDistribution.settings")
setup()
from main.models import Task, TaskAssignment, Employee, EmployeeConfirmation, Issue, EmployeeRating, Notes, \
    TaskCompletionRequest
from django.db import IntegrityError

bot_token = '6262154198:AAFR2VQZpzBaojlaGyEFo-jK4GbpDCTz91E'
bot = telebot.TeleBot(bot_token)

states = {}
failed_tasks = []
reminders = {}
timers = {}


@bot.message_handler(commands=['start'])
def handle_start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_registration = types.KeyboardButton("Регистрация")
    item_authorization = types.KeyboardButton("Авторизация")
    markup.add(item_registration, item_authorization)

    bot.send_message(message.chat.id, "Добро пожаловать!", reply_markup=markup)


# Обработчик для команды '/start'
@bot.message_handler(commands=['start'])
def handle_start(message):
    # Создание кастомной клавиатуры с опциями регистрации и авторизации
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_registration = types.KeyboardButton("Регистрация")
    item_authorization = types.KeyboardButton("Авторизация")
    markup.add(item_registration, item_authorization)

    # Отправка приветственного сообщения с кастомной клавиатурой
    bot.send_message(message.chat.id, "Добро пожаловать!", reply_markup=markup)


# Обработчик для опции 'Регистрация'
@bot.message_handler(func=lambda message: message.text == "Регистрация")
def handle_registration(message):
    bot.send_message(message.chat.id, "Введите логин:")
    bot.register_next_step_handler(message, process_login)


# Функция для обработки логина пользователя в процессе регистрации
def process_login(message):
    login = message.text
    # Проверка, содержит ли логин только буквы и цифры
    if not login.isalnum():
        bot.send_message(message.chat.id, "Логин не должен содержать специальных символов. Попробуйте еще раз.")
        handle_registration(message)
        return

    # Сохранение логина в состоянии пользователя
    states[message.from_user.id] = {'login': login}

    bot.send_message(message.chat.id, "Введите пароль:")
    bot.register_next_step_handler(message, process_password)


# Функция для обработки пароля пользователя в процессе регистрации
def process_password(message):
    login = states[message.from_user.id]['login']
    password = message.text

    # Проверка требований к паролю
    if not (
            any(c.islower() for c in password)
            and any(c.isupper() for c in password)
            and any(c.isdigit() for c in password)
            and any(c.isascii() and not c.isalnum() for c in password)
            and len(password) >= 8
    ):
        bot.send_message(
            message.chat.id,
            "Пароль должен содержать не менее 8 символов, включая маленькие и большие буквы, цифры и специальные "
            "символы.",
        )
        handle_registration(message)
        return

    # Отправка сообщения об успешной регистрации и сохранение пользователя в базе данных
    # bot.send_message(message.chat.id, "Ваш аккаунт ожидает подтверждения. Можете попробовать авторизироваться.")
    try:
        bot.send_message(message.chat.id, "Введите ваше имя:")
        bot.register_next_step_handler(message, process_name, login, password)
    except IntegrityError:
        print("Ошибка IntegrityError.")
        pass
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        pass


def process_name(message, login, password):
    name = message.text

    # Повторите процесс для фамилии
    bot.send_message(message.chat.id, "Введите вашу фамилию:")
    bot.register_next_step_handler(message, process_surname, login, password, name)


def process_surname(message, login, password, name):
    surname = message.text

    # Повторите процесс для отчества
    bot.send_message(message.chat.id, "Введите ваше отчество (если есть):")
    bot.register_next_step_handler(message, process_patronymic, login, password, name, surname)


def process_patronymic(message, login, password, name, surname):
    patronymic = message.text

    # Повторите процесс для почты
    bot.send_message(message.chat.id, "Введите вашу почту:")
    bot.register_next_step_handler(message, process_email, login, password, name, surname, patronymic)


def process_email(message, login, password, name, surname, patronymic):
    email = message.text

    # Сохранение информации в базу данных
    bot.send_message(message.chat.id, "Ваш аккаунт ожидает подтверждения. Можете попробовать авторизироваться.")

    try:
        employee, created = Employee.objects.get_or_create(
            login=login, password=password, name=name, surname=surname, patronymic=patronymic, email=email
        )
        employee.save()
    except IntegrityError:
        print("Ошибка IntegrityError.")
        pass
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        pass


# Функция для отправки уведомления конкретному сотруднику
def send_notification(employee, message):
    print(f"Уведомление для сотрудника {employee}: {message}")


# Функция для отправки уведомлений о предстоящих сроках задач
def send_deadline_notifications():
    now = timezone.now()
    deadline_alert_time = now + timedelta(days=1)
    # Фильтрация задач с предупреждением о дедлайне на следующий день
    tasks_to_notify = Task.objects.filter(deadline=deadline_alert_time.date())
    for task in tasks_to_notify:
        # Проверка наличия исполнителя и отправка уведомления
        if task.employee:
            message_text = f"Завтра дедлайн по задаче: {task.title}"
            # Если задача срочная, добавляем соответствующее уточнение
            if task.status == 'срочно':
                message_text += " (срочно)"
            # Создаем кастомную клавиатуру с кнопкой "Отказаться"
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton(text="Отказаться", callback_data=f"refuse_{task.id}"))

            send_notification(task.employee, message_text)

            # Обновление поля 'notifications_sent' для задачи
            task.notifications_sent += 1
            task.save()


# Обработчик для опции 'Авторизация'
@bot.message_handler(func=lambda message: message.text == "Авторизация")
def handle_authorization(message):
    bot.send_message(message.chat.id, "Введите логин:")
    bot.register_next_step_handler(message, process_login_authorization)
    # Отправка уведомлений о дедлайнах после обработки авторизации
    send_deadline_notifications()


# Функция для обработки логина пользователя в процессе авторизации
def process_login_authorization(message):
    login = message.text
    bot.send_message(message.chat.id, "Введите пароль:")
    bot.register_next_step_handler(message, process_password_authorization, login)
    # Отправка уведомлений о дедлайнах после обработки авторизации
    send_deadline_notifications()


# Функция для обработки пароля пользователя в процессе авторизации
def process_password_authorization(message, login):
    # Отправка уведомлений о дедлайнах после обработки авторизации
    send_deadline_notifications()
    password = message.text
    employee = Employee.objects.filter(login=login, password=password).first()

    if employee:
        if employee.is_blocked:
            bot.send_message(message.chat.id, "Вы заблокированы. Обратитесь к администратору.")
            return

        # Сохранение логина в состоянии пользователя
        states[message.from_user.id] = {'login': login}
        confirmation = EmployeeConfirmation.objects.filter(Q(employee=employee)).first()

        if confirmation:
            # Создание кастомной клавиатуры для авторизованных пользователей
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item_get_tasks = types.KeyboardButton("Получение задач")
            item_task_list = types.KeyboardButton("Задачи в процессе выполнения")
            item_calendar = types.KeyboardButton("Показать календарь")
            item_export_excel = types.KeyboardButton("Выгрузить в Excel")
            item_logout = types.KeyboardButton("Выйти из аккаунта")
            markup.add(item_get_tasks, item_task_list, item_calendar, item_export_excel, item_logout)

            # Отправка сообщения об успешной авторизации с кастомной клавиатурой
            bot.send_message(message.chat.id, "Авторизация успешна! Выберите действие:", reply_markup=markup)
            bot.clear_step_handler(message)
        else:
            bot.send_message(message.chat.id, "Не удалось войти в аккаунт! Возможно, его еще не подтвердили...")
            handle_authorization(message)
    else:
        bot.send_message(message.chat.id, "Неверный логин или пароль. Попробуйте еще раз.")


# Обработчик для опции 'Показать календарь'
@bot.message_handler(func=lambda message: message.text == "Показать календарь")
def handle_show_calendar(message):
    show_calendar(message)


# Функция для отображения календаря для выбора даты и времени
def show_calendar(message):
    now = datetime.now()
    markup = types.InlineKeyboardMarkup(row_width=7)

    year = now.year
    month = now.month
    cal = calendar.monthcalendar(year, month)

    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(types.InlineKeyboardButton(" ", callback_data="ignore"))
            else:
                row.append(types.InlineKeyboardButton(str(day), callback_data=f"day_{day}"))

        markup.add(*row)

    bot.send_message(message.chat.id, "Выберите день:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith('Отказаться'))
def handle_refuse_task(call):
    task_id = int(call.data.split('_')[1])
    task = Task.objects.get(id=task_id)

    # Сохраняем информацию о невыполненной задаче
    failed_task_info = f"{call.from_user.first_name} {call.from_user.last_name} отказался от задачи: {task.title}"
    failed_tasks.append(failed_task_info)

    # Обновляем статус задачи
    task.status = 'свободна'
    task.save()

    # Отправляем уведомление об отказе
    bot.send_message(call.message.chat.id, f"Вы отказались от задачи: {task.title}")


@bot.message_handler(func=lambda message: message.text == "Выгрузить в Excel")
def export_to_excel(message):
    employee_list = Employee.objects.all()
    rating_list = EmployeeRating.objects.all()
    task_assignments = TaskAssignment.objects.all()

    wb = openpyxl.Workbook()

    employee_sheet = wb.create_sheet(title="Сотрудники")
    employee_sheet.append(["Имя", "Фамилия", "Отчество", "Логин", "Почта", "Должность", "Зарплата", "Отдел"])
    for employee in employee_list:
        employee_sheet.append([
            employee.name if employee.name else "",
            employee.surname if employee.surname else "",
            employee.patronymic if employee.patronymic else "",
            employee.login if employee.login else "",
            employee.email if employee.email else "",
            employee.position if employee.position else "",
            employee.salary if employee.salary else "",
            employee.department.name if employee.department else ""
        ])

    rating_sheet = wb.create_sheet(title="Рейтинг")
    rating_sheet.append(["Сотрудник", "Рейтинг", "Заблокирован"])
    for rating_entry in rating_list:
        rating_sheet.append([
            str(rating_entry.employee),
            rating_entry.rating,
            rating_entry.is_blocked
        ])

    failed_tasks_sheet = wb.create_sheet(title="Невыполненные задачи")
    failed_tasks_sheet.append(["Информация о невыполненных задачах"])
    for failed_task_info in failed_tasks:
        failed_tasks_sheet.append([failed_task_info])

    # Лист "Мои задачи"
    tasks_sheet = wb.create_sheet(title="Мои задачи")
    tasks_sheet.append(["Дедлайн", "Название задачи", "Статус"])
    for task_assignment in task_assignments:
        task = task_assignment.task
        tasks_sheet.append([
            task.deadline.strftime("%Y-%m-%d %H:%M:%S"),
            task.title,
            task.status
        ])

    # Сохранение файла
    file_path = "employee_data.xlsx"
    wb.save(file_path)

    # Отправка файла в чат
    with open(file_path, "rb") as file:
        bot.send_document(message.chat.id, file)

    # Опционально: удаление файла после отправки (если нужно)
    os.remove(file_path)


# Обработчик выбора даты в календаре
@bot.callback_query_handler(func=lambda call: call.data.startswith("day_"))
def handle_day_selection(call):
    selected_day = int(call.data.split("_")[1])
    selected_date = datetime.now().replace(day=selected_day, hour=0, minute=0, second=0, microsecond=0)

    # Запрашиваем пользователя ввести время напоминания
    bot.send_message(call.message.chat.id, f"Вы выбрали {selected_date.strftime('%Y-%m-%d')}. Введите время (чч:мм):")
    bot.register_next_step_handler(call.message, save_time, selected_date)


# Обработчик ввода времени напоминания
def save_time(message, selected_date):
    try:
        # Преобразование строки времени пользователя в объект datetime
        time_str = message.text
        selected_time = datetime.strptime(time_str, '%H:%M')

        # Комбинирование выбранной даты и времени
        selected_datetime = selected_date + timedelta(hours=selected_time.hour, minutes=selected_time.minute)

        # Запрос пользователя на текст напоминания
        bot.send_message(message.chat.id, "Введите текст напоминания:")
        bot.register_next_step_handler(message, save_note, selected_datetime)
    except ValueError:
        bot.send_message(message.chat.id, "Некорректный формат времени. Попробуйте еще раз.")


# Обработчик ввода текста напоминания и сохранение в базе данных
def save_note(message, selected_datetime):
    note_text = message.text
    user_id = message.from_user.id

    # Получаем текущего сотрудника
    login = states[user_id]['login']
    current_employee = Employee.objects.filter(login=login).first()

    # Сохраняем напоминание в базе данных
    reminder = Notes(employee=current_employee, datetime=selected_datetime, text=note_text)
    reminder.save()

    bot.send_message(message.chat.id, "Напоминание успешно сохранено!")


# Обработчик для просмотра всех напоминаний на выбранный день
@bot.message_handler(commands=['view_reminders'])
def view_reminders(message):
    user_id = message.from_user.id
    if user_id in reminders:
        reminders_for_user = reminders[user_id]
        formatted_reminders = [
            f"{reminder['datetime'].strftime('%Y-%m-%d %H:%M')}: {reminder['text']}"
            for reminder in reminders_for_user
        ]
        bot.send_message(message.chat.id, "Ваши напоминания на этот день:\n" + "\n".join(formatted_reminders))
    else:
        bot.send_message(message.chat.id, "На выбранный день у вас нет напоминаний.")


# Функция для получения задач для определенного отдела
def get_department_tasks(department):
    tasks = Task.objects.filter(department=department)
    return tasks


# Обработчик для опции 'Получение задач'
@bot.message_handler(func=lambda message: message.text == "Получение задач")
def handle_get_tasks(message):
    # Отправка уведомлений о дедлайнах перед обработкой получения задач
    send_deadline_notifications()
    login = states[message.from_user.id]['login']
    employee = Employee.objects.get(login=login)

    if employee.department is None:
        bot.send_message(message.chat.id, "В данный момент нет данных о вашем отделе. Обратитесь к администратору.")
        return

    # Получение задач для отдела текущего пользователя
    department_tasks = get_department_tasks(employee.department)

    if department_tasks:
        markup = types.InlineKeyboardMarkup(row_width=1)

        print(department_tasks)
        for task in department_tasks:
            if task.status != "В процессе":
                # Создание кнопки для каждой задачи в отделе
                task_button = types.InlineKeyboardButton(text=task.title, callback_data=f"task_{task.id}")
                markup.add(task_button)

        # Отправка сообщения с выбором задачи из списка
        bot.send_message(message.chat.id, "Выберите задачу из списка:", reply_markup=markup)


# Регистрация обработчика для выбора задачи
@bot.callback_query_handler(func=lambda call: call.data.startswith("task_"))
def process_task_selection(query):
    try:
        task_id = int(query.data.split("_")[1])
        selected_task = Task.objects.get(id=task_id)
        user_id = query.from_user.id
        login = states[user_id]['login']
        employee = Employee.objects.filter(login=login).first()

        if selected_task and employee:
            if selected_task.difficulty < employee.rating:
                bot.send_message(query.message.chat.id, "У вас недостаточный уровень для выполнения этой задачи.")
                return

            selected_task.status = "В процессе"
            selected_task.save()

            task_assignment, created = TaskAssignment.objects.get_or_create(task=selected_task)
            if created:
                task_assignment.employee = employee
                task_assignment.save()

            if created:
                bot.send_message(query.message.chat.id,
                                 f"Выбранная вами задача '{selected_task.title}' зарезервирована за вами.")
            else:
                bot.send_message(query.message.chat.id, f"Вы уже взяли задачу '{selected_task.title}'.")
        else:
            pass
    except (IndexError, ValueError, Task.DoesNotExist):
        pass


# Обработчик для опции 'Список задач в процессе выполнения'
@bot.message_handler(func=lambda message: message.text == "Задачи в процессе выполнения")
def handle_all_task_list(message):
    all_tasks = TaskAssignment.objects.all()
    # Отправка списка задач с использованием кастомной клавиатуры
    send_task_list(message, all_tasks, is_inline=True)


# Функция для отправки списка задач
def send_task_list(message, task_assignments, is_inline=False):
    login = states[message.from_user.id]['login']
    employee = Employee.objects.filter(login=login).first()

    if task_assignments:
        markup = types.InlineKeyboardMarkup(row_width=1) if is_inline else types.ReplyKeyboardMarkup(
            resize_keyboard=True)

        for task_assignment in task_assignments:
            task = task_assignment.task
            if task_assignment.employee == employee:
                # Используйте strftime для форматирования дедлайна
                deadline_str = task.deadline.strftime('%Y-%m-%d %H:%M:%S')
                # Включите дедлайн в название задачи при выводе списка
                task_buttons = [
                    types.InlineKeyboardButton(text=f"{task.title} ({deadline_str})",
                                               callback_data=f"task_{task.id}"),
                    types.InlineKeyboardButton(text="Сдвинуть срок", callback_data=f"postpone_{task.id}"),
                    types.InlineKeyboardButton(text="Отказ от задачи", callback_data=f"reject_{task.id}"),
                    types.InlineKeyboardButton(text="Выполнил задачу", callback_data=f"complete_{task.id}")
                ]
                markup.add(*task_buttons)

        if is_inline:
            try:
                bot.send_message(message.chat.id, "Ваши задачи:", reply_markup=markup)
            except Exception as e:
                print(f"Error sending message: {e}")
        else:
            bot.send_message(message.chat.id, "Выберите задачу для завершения:", reply_markup=markup)
            bot.register_next_step_handler(message, complete_task_handler)
    else:
        bot.send_message(message.chat.id, "Нет задач в процессе выполнения.")



# Обработчик обратного вызова для команды 'postpone_'
@bot.callback_query_handler(func=lambda call: 'postpone_' in call.data)
def move_deadline_handler(call):
    # Запрос причины смены срока
    bot.send_message(call.message.chat.id, "Введите причину смены срока:")
    bot.register_next_step_handler(call.message, ask_for_reason, call.data)


# Функция для запроса причины и нового срока
def ask_for_reason(message, callback_data):
    # Сохраняем причину в глобальной переменной
    task_id = int(callback_data.split('_')[1])
    reason = message.text
    states[message.from_user.id]['reason'] = reason

    # Запрашиваем новый срок с использованием Inline Keyboard
    markup = types.InlineKeyboardMarkup(row_width=1)
    today = datetime.now()
    for i in range(1, 8):
        date_option = today + timedelta(days=i)
        callback_option = f"confirm_{task_id}_{date_option.strftime('%Y-%m-%d')}"
        button = types.InlineKeyboardButton(text=date_option.strftime('%Y-%m-%d'), callback_data=callback_option)
        markup.add(button)

    bot.send_message(message.chat.id, "Выберите новый срок:", reply_markup=markup)


# Обработчик подтверждения смены срока
@bot.callback_query_handler(func=lambda call: 'confirm_' in call.data)
def confirm_postpone_handler(call):
    task_id = int(call.data.split('_')[1])
    new_deadline = datetime.strptime(call.data.split('_')[2], '%Y-%m-%d')

    reason = states[call.from_user.id]['reason']
    login = states[call.from_user.id]['login']
    current_employee = Employee.objects.filter(login=login).first()
    selected_task = Task.objects.filter(id=task_id).first()

    if selected_task and current_employee:
        # Обновление срока задачи
        selected_task.deadline = new_deadline
        selected_task.save()

        # Создание Issue с указанием даты
        issue = Issue(task=selected_task, employee=current_employee, status=reason, date=new_deadline)
        issue.save()

        bot.send_message(call.message.chat.id,
                         f"Запрос отправлен на обработку. Ожидайте ответа")
    else:
        error_message = f"Ошибка при обработке запроса. selected_task: {selected_task}, current_employee: {current_employee}"
        bot.send_message(call.message.chat.id, error_message)


@bot.callback_query_handler(func=lambda call: 'task_' in call.data)
def move_deadline_handler():
    pass


# Обработчик обратного вызова для команды 'reject_'
@bot.callback_query_handler(func=lambda call: 'reject_' in call.data)
def reject_task_handler(call):
    task_id = int(call.data.split('_')[1])
    login = states[call.from_user.id]['login']
    current_employee = Employee.objects.filter(login=login).first()
    selected_task = Task.objects.filter(id=task_id).first()

    if selected_task and current_employee:
        selected_task.status = 'Свободна'
        current_employee.rating -= 1
        selected_task.save()
        current_employee.save()

        task_assignment = TaskAssignment.objects.filter(task=selected_task, employee=current_employee).first()
        if task_assignment:
            task_assignment.delete()

        bot.send_message(call.message.chat.id, "Вы отказались от задачи.")
    else:
        bot.send_message(call.message.chat.id, "Произошла ошибка при обработке запроса.")


# Обработчик обратного вызова для команды 'complete_'
@bot.callback_query_handler(func=lambda call: 'complete_' in call.data)
def complete_task_handler(call):
    task_id = int(call.data.split('_')[1])
    login = states[call.from_user.id]['login']
    current_employee = Employee.objects.filter(login=login).first()
    selected_task = Task.objects.filter(id=task_id).first()

    if selected_task and current_employee:
        completion_request = TaskCompletionRequest(task=selected_task, employee=current_employee)
        completion_request.save()

        bot.send_message(call.message.chat.id, "Заявка на завершение задачи направлена на модерацию.")

    else:
        bot.send_message(call.message.chat.id, "Произошла ошибка при обработке запроса.")


# Обработчик для опции 'Выйти из аккаунта'
@bot.message_handler(func=lambda message: message.text == "Выйти из аккаунта")
def handle_logout(message):

    # Удаление кастомной клавиатуры и переход в начальное состояние
    markup = types.ReplyKeyboardRemove(selective=False)
    bot.send_message(message.chat.id, "Вы успешно вышли из аккаунта.", reply_markup=markup)

    # Создание начальной клавиатуры для выбора действия
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_registration = types.KeyboardButton("Регистрация")
    item_authorization = types.KeyboardButton("Авторизация")
    markup.add(item_registration, item_authorization)

    # Отправка сообщения с начальной клавиатурой
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)

    global states, failed_tasks, reminders, timers

    states.clear()
    failed_tasks.clear()
    reminders.clear()
    timers.clear()

# Запуск бота
if __name__ == "__main__":
    bot.polling(none_stop=True)
